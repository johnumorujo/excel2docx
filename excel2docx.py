#!/usr/bin/python3
from openpyxl import load_workbook
from tqdm import tqdm
import sys

# Check if we are running this on windows platform
is_windows = sys.platform.startswith('win')
if is_windows:
    # Windows deserves coloring too :D
    G = '\033[92m'  # green
    Y = '\033[93m'  # yellow
    B = '\033[94m'  # blue
    R = '\033[91m'  # red
    W = '\033[0m'  # white
    try:
        import win_unicode_console
        import colorama

        win_unicode_console.enable()
        colorama.init()
        # Now the unicode will work ^_^
    except 1 == 2:
        print("[!] Error: Coloring libraries not installed, no coloring will be used [Check the readme]")
        G = Y = B = R = W = ''


else:
    G = '\033[92m'  # green
    Y = '\033[93m'  # yellow
    B = '\033[94m'  # blue
    R = '\033[91m'  # red
    W = '\033[0m'  # white


def no_color():
    global G, Y, B, R, W
    G = Y = B = R = W = ''


def banner():
    print("""%s


EEEEEEEEEEEEEEEEEEEEEEXXXXXXX       XXXXXXX        CCCCCCCCCCCCCEEEEEEEEEEEEEEEEEEEEEELLLLLLLLLLL                       
E::::::::::::::::::::EX:::::X       X:::::X     CCC::::::::::::CE::::::::::::::::::::EL:::::::::L                       
E::::::::::::::::::::EX:::::X       X:::::X   CC:::::::::::::::CE::::::::::::::::::::EL:::::::::L                       
EE::::::EEEEEEEEE::::EX::::::X     X::::::X  C:::::CCCCCCCC::::CEE::::::EEEEEEEEE::::ELL:::::::LL                       
  E:::::E       EEEEEEXXX:::::X   X:::::XXX C:::::C       CCCCCC  E:::::E       EEEEEE  L:::::L                         
  E:::::E                X:::::X X:::::X   C:::::C                E:::::E               L:::::L                         
  E::::::EEEEEEEEEE       X:::::X:::::X    C:::::C                E::::::EEEEEEEEEE     L:::::L                         
  E:::::::::::::::E        X:::::::::X     C:::::C                E:::::::::::::::E     L:::::L                         
  E:::::::::::::::E        X:::::::::X     C:::::C                E:::::::::::::::E     L:::::L                         
  E::::::EEEEEEEEEE       X:::::X:::::X    C:::::C                E::::::EEEEEEEEEE     L:::::L                         
  E:::::E                X:::::X X:::::X   C:::::C                E:::::E               L:::::L                         
  E:::::E       EEEEEEXXX:::::X   X:::::XXX C:::::C       CCCCCC  E:::::E       EEEEEE  L:::::L         LLLLLL          
EE::::::EEEEEEEE:::::EX::::::X     X::::::X  C:::::CCCCCCCC::::CEE::::::EEEEEEEE:::::ELL:::::::LLLLLLLLL:::::L          
E::::::::::::::::::::EX:::::X       X:::::X   CC:::::::::::::::CE::::::::::::::::::::EL::::::::::::::::::::::L          
E::::::::::::::::::::EX:::::X       X:::::X     CCC::::::::::::CE::::::::::::::::::::EL::::::::::::::::::::::L          
EEEEEEEEEEEEEEEEEEEEEEXXXXXXX      TTTTTTTTTTTTTTTTTTTTTTTCCCCCOOOOOOOOOEEEEEEEEEEEEEELLLLLLLLLLLLLLLLLLLLLLLL          
                                   T:::::::::::::::::::::T   OO:::::::::OO                                              
                                   T:::::::::::::::::::::T OO:::::::::::::OO                                            
                                   T:::::TT:::::::TT:::::TO:::::::OOO:::::::O                                           
                                   TTTTTT  T:::::T  TTTTTTO::::::O   O::::::O                                           
                                           T:::::T        O:::::O     O:::::O                                           
                                           T:::::T        O:::::O     O:::::O                                           
                                           T:::::T        O:::::O     O:::::O                                           
                                           T:::::T        O:::::O     O:::::O                                           
                                           T:::::T        O:::::O     O:::::O                                           
                                           T:::::T        O:::::O     O:::::O                                           
                                           T:::::T        O::::::O   O::::::O                                           
                                         TT:::::::TT      O:::::::OOO:::::::O                                           
                                         T:::::::::T       OO:::::::::::::OO                                            
                                         T:::::::::T         OO:::::::::OO                                              
               DDDDDDDDDDDDD             OOOOOOOOOTT           CCCCCCCCCCCCCXXXXXXX       XXXXXXX                       
               D::::::::::::DDD        OO:::::::::OO        CCC::::::::::::CX:::::X       X:::::X                       
               D:::::::::::::::DD    OO:::::::::::::OO    CC:::::::::::::::CX:::::X       X:::::X                       
               DDD:::::DDDDD:::::D  O:::::::OOO:::::::O  C:::::CCCCCCCC::::CX::::::X     X::::::X                       
                 D:::::D    D:::::D O::::::O   O::::::O C:::::C       CCCCCCXXX:::::X   X:::::XXX                       
                 D:::::D     D:::::DO:::::O     O:::::OC:::::C                 X:::::X X:::::X                          
                 D:::::D     D:::::DO:::::O     O:::::OC:::::C                  X:::::X:::::X                           
                 D:::::D     D:::::DO:::::O     O:::::OC:::::C                   X:::::::::X                            
                 D:::::D     D:::::DO:::::O     O:::::OC:::::C                   X:::::::::X                            
                 D:::::D     D:::::DO:::::O     O:::::OC:::::C                  X:::::X:::::X                           
                 D:::::D     D:::::DO:::::O     O:::::OC:::::C                 X:::::X X:::::X                          
                 D:::::D    D:::::D O::::::O   O::::::O C:::::C       CCCCCCXXX:::::X   X:::::XXX                       
               DDD:::::DDDDD:::::D  O:::::::OOO:::::::O  C:::::CCCCCCCC::::CX::::::X     X::::::X                       
               D:::::::::::::::DD    OO:::::::::::::OO    CC:::::::::::::::CX:::::X       X:::::X                       
               D::::::::::::DDD        OO:::::::::OO        CCC::::::::::::CX:::::X       X:::::X                       
               DDDDDDDDDDDDD             OOOOOOOOO             CCCCCCCCCCCCCXXXXXXX       XXXXXXX                      






                                                                                                                        %s%s
                # Coded By Umoru John - github-- /johnumorujo | instagram-- @johnumorujo
    """ % (G, W, Y))


banner()
print(
    W + """Usage: 'excel2docx' + 'name of excel file' + 'Sheet1(or the name of the spreadsheet)***case sensitive' + 'name of output file'""" + B)
print("""edit #output area to suit your needs
""" + G)
# system arguments

note = sys.argv[1]
output = sys.argv[3]
wb = load_workbook(str(note))
sheet = sys.argv[2]
# print(wb.sheetnames)
ws = wb[f'{str(sheet)}']
# select min row and column here!
minrow, maxrow, mincol, maxcol = 2, ws.max_row, 1, 6
rows = ws.iter_rows(minrow, maxrow, mincol, maxcol)
# you can comment this out and also remove it from #OUTPUT
numbe = 1
newname = f'{str(output)}.docx'
fyle = open(newname, "a")
print(f"converting {str(note)} to {str(output)}.docx" + W)

# output
for a, b, c, d, e, f in tqdm(rows):
    q = a.value
    a = b.value
    b = c.value
    c = d.value
    d = e.value
    ans = f.value
    # you can start edit here
    fyle.write(f"""{numbe}: {q}
(a): {a}
(b): {b}
(c): {c}
(d): {d}
Answer: {ans}

""")
    # you should stop edit here
    numbe += 1
fyle.close()
print(f"done converting {str(note)} to {str(output)}.docx")
